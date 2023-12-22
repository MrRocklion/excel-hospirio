import openpyxl
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image
from datetime import datetime
from firebase_admin import storage
from uuid_extensions import  uuid7str
from os import remove
## 
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
import os
#agregamos las credenciales de firebase
cred = credentials.Certificate('credentials.json')
app_firebase = firebase_admin.initialize_app(cred,
                                             {
    'storageBucket': 'app-mantenimiento-91156.appspot.com'
})
db = firestore.client()
bucket = storage.bucket()
# funciona workbook

imagen = Image('logo_hospital.jpg')
# Cambiar el ancho y alto de la imagen
meses = ['','Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre','']
def generarExcel(data):
    workbook = openpyxl.Workbook()
    docs = db.collection("ingreso").stream()
    docs_mantenimientos = []
    for doc in docs:
        aux = doc.to_dict()
        if data['departamento'] != 1000 and data['tipo'] != 0:
            print(aux['codigo'])
            if data['tipo'] == aux['tipo_equipo']['codigo'] and data['departamento'] == aux['departamento']['codigo']:
                if len(aux['mantenimientos']) > 0 and aux['situacion'] == 'Activo':
                    for i in aux['mantenimientos']:
                        docs_mantenimientos.append(i)
        elif data['departamento'] != 1000 and data['tipo'] == 0 :
            print(aux['codigo'])
            if data['departamento'] == aux['departamento']['codigo']:
                if len(aux['mantenimientos']) > 0 and aux['situacion'] == 'Activo':
                    for i in aux['mantenimientos']:
                        docs_mantenimientos.append(i)
        elif data['departamento'] == 1000 and data['tipo'] != 0 :
            print(aux['codigo'])
            if data['tipo'] == aux['tipo_equipo']['codigo']:
                if len(aux['mantenimientos']) > 0 and aux['situacion'] == 'Activo':
                    for i in aux['mantenimientos']:
                        docs_mantenimientos.append(i)
    
        else :
            if len(aux['mantenimientos']) > 0 and aux['situacion'] == 'Activo':
                for i in aux['mantenimientos']:
                    docs_mantenimientos.append(i)


    print(len(docs_mantenimientos))
    aux_mes = data['month']
    aux_year = data['year']
    man_filtrados = []
    #listamos todos los mantenimientos
    for man in docs_mantenimientos:
        aux_fecha = man['start']
        fecha_obj = datetime.strptime(aux_fecha, '%m/%d/%Y, %I:%M:%S %p')
        mes = fecha_obj.month
        year = fecha_obj.year
        print(f"comparamos {mes} con el que vino por parametro {aux_mes}")
        if mes == aux_mes and year==aux_year:
            man_filtrados.append(man)


    imagen.width = 170  # Reemplaza 200 con el ancho deseado en píxeles
    imagen.height = 50 
    borde_negro_grueso = Border(left=Side(style='thin'), 
                                right=Side(style='thin'), 
                                top=Side(style='thin'), 
                                bottom=Side(style='thin'))
    centrar_texto = Alignment(horizontal='center', vertical='center')
    sheet = workbook.active

    sheet.add_image(imagen, 'A1')
    sheet['A1'].border = borde_negro_grueso
    sheet['A1'].alignment = centrar_texto
    sheet['D1'] = 'Reporte de mantenimientos preventivos mensual'
    sheet['D1'].border = borde_negro_grueso
    sheet['D1'].alignment = centrar_texto
    sheet['I1'] = 'ING-FO-04'
    sheet['I1'].border = borde_negro_grueso
    sheet['I1'].alignment = centrar_texto
    sheet['I2'] = 'Revision: 00'
    sheet['I2'].border = borde_negro_grueso
    sheet['I2'].alignment = centrar_texto
    sheet['I3'] = data['date']
    sheet['I3'].border = borde_negro_grueso
    sheet['I3'].alignment = centrar_texto
    #sheet.merge_cells('A1:C1')
    sheet['A5'].border = borde_negro_grueso
    sheet['A5'] = f'Año: {aux_year} Mes: {meses[aux_mes]}'
    sheet.merge_cells('I1:K1')
    sheet.merge_cells('I2:K2')
    sheet.merge_cells('I3:K3')
    sheet.merge_cells('D1:H3')
    sheet.merge_cells('A1:C3')
    sheet.merge_cells('A4:K4')
    sheet.merge_cells('A5:K5')


    # encabezados de los datos
    sheet['A6'] = 'N.'
    sheet['A6'].border = borde_negro_grueso
    sheet['A6'].alignment = centrar_texto
    sheet.merge_cells('A6:A7')
    sheet['B6'] = 'Nombre del Equipo'
    sheet['B6'].border = borde_negro_grueso
    sheet['B6'].alignment = centrar_texto
    sheet.merge_cells('B6:E7')
    sheet['F6'] = 'Codigo del equipo'
    sheet['F6'].border = borde_negro_grueso
    sheet['F6'].alignment = centrar_texto
    sheet.merge_cells('F6:H7')
    sheet['I6'] = 'Cumplimiento'
    sheet['I6'].border = borde_negro_grueso
    sheet['I6'].alignment = centrar_texto
    sheet.merge_cells('I6:K6')
    sheet['I7'] = 'SI'
    sheet['I7'].border = borde_negro_grueso
    sheet['I7'].alignment = centrar_texto
    sheet['J7'] = 'NO'
    sheet['J7'].border = borde_negro_grueso
    sheet['J7'].alignment = centrar_texto
    sheet['K7'] = 'Observaciones'
    sheet['K7'].border = borde_negro_grueso
    sheet['K7'].alignment = centrar_texto
    sheet.column_dimensions['I'].width = 5
    sheet.column_dimensions['J'].width = 5
    sheet.column_dimensions['K'].width = 15
    sheet.row_dimensions[4].height = 20

    counter = 1
    for i in man_filtrados:
        print(i)
        sheet[f'A{counter+7}'] = counter
        sheet[f'A{counter+7}'].border = borde_negro_grueso
        sheet[f'B{counter+7}'] = i['title']
        sheet[f'B{counter+7}'].border = borde_negro_grueso
        sheet[f'F{counter+7}'] = i['codigo_equipo']
        sheet[f'F{counter+7}'].border = borde_negro_grueso
        sheet[f'K{counter+7}'] = 'ninguna'
        sheet[f'K{counter+7}'].border = borde_negro_grueso
        if i['verificacion'] == True:
            sheet[f'I{counter+7}'] = 'x'
            
        else:
            sheet[f'J{counter+7}'] = 'x'
        sheet[f'I{counter+7}'].border = borde_negro_grueso
        sheet[f'J{counter+7}'].border = borde_negro_grueso
        sheet.merge_cells(f'B{counter+7}:E{counter+7}')
        sheet.merge_cells(f'F{counter+7}:H{counter+7}')
        counter +=1

    #sheet.merge_cells('D1:H1')
    #sheet.merge_cells('D1:D3')
    # Guardar el libro de trabajo en un archivo
    sheet.row_dimensions[4].height = 6 # 1 pulgada
    id_file = uuid7str()
    workbook.save(f'{id_file}.xlsx')
    blob = bucket.blob(f'mantenimientos/{id_file}.xlsx')
    blob.upload_from_filename(f'{id_file}.xlsx')
    blob.make_public()

    print("your file url", blob.public_url)
    print("Archivo Excel generado con éxito.")
    return blob.public_url
